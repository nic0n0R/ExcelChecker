#
# Name: Excel Files comparator
# Author: nic0n0R
# Version: 0.1
#


import openpyxl
from openpyxl import styles
import xlrd
import tkinter
from tkinter import filedialog as fd
from tkinter import messagebox
import os

"""
1. Закидываем два файла
2. Считываем нужные данные
3. Сравниваем отклонения
4. Форматирование


###
отклонение - 17 столбец
###
"""
out_path = None
path1 = ''
path2 = ''
wbf1 = None
wbf2 = None
ws1 = None
ws2 = None
names_file1 = []
offset_file1 = []
names_file2 = []
offset_file2 = []
count1 = []
count2 = []
first_dict = {}
second_dict = {}
comparedYellow = []
comparedGreen = []
comparedRed = []
first_file = []
second_file = []


def clear_vars():
    global out_path, path1, path2, wbf1, wbf2, ws1, ws2, names_file1, names_file2, offset_file1, offset_file2, count1, \
        count2, first_file, second_file, comparedYellow, comparedGreen, comparedRed, first_dict, second_dict
    out_path = None
    path1 = ''
    path2 = ''
    wbf1 = None
    wbf2 = None
    ws1 = None
    ws2 = None
    names_file1 = []
    offset_file1 = []
    names_file2 = []
    offset_file2 = []
    count1 = []
    count2 = []
    first_dict = {}
    second_dict = {}
    comparedYellow = []
    comparedGreen = []
    comparedRed = []
    first_file = []
    second_file = []


def download_info(filename, names, offsets, count):
    wb = xlrd.open_workbook(filename)
    sheet = wb.sheet_by_index(0)
    row_number = sheet.nrows

    if row_number > 0:
        for row in range(row_number):
            names.append(str(sheet.row(row)[2]).replace('text:', '').replace('\'', ''))
            offsets.append(str(sheet.row(row)[17]).replace('text:', '').replace('number:', ''))
            count.append(
                str(sheet.row(row)[0]).replace('text:', '').replace('number:', '').replace('№', '').replace('empty:',
                                                                                                            '').replace(
                    '\'', ''))


def case_finding(value, work_sheet):
    rows = work_sheet.nrows
    cols = work_sheet.ncols
    tmp = 0
    for row in rows:
        if (work_sheet.row(row)[2]) == value:
            tmp = row
            return 'R{}'.format(tmp + 1)


"""
Yellow - Не найдено совпадений позиции в двух файлах
Green  - Отколенение в обоих файлах одинаковое
Red    - Отклонения в файлах отличаются 
"""


def formatting(obj, type_formatting):
    global wbf1, wbf2, ws1, ws2
    wbf1 = openpyxl.load_workbook('E:\\8_apr.xlsx')
    wbf2 = openpyxl.load_workbook('E:\\27_yanvarya.xlsx')
    ws1 = wbf1['Лист1']
    ws2 = wbf2['Лист1']
    yellowFill = styles.PatternFill(fill_type='solid', start_color=styles.colors.YELLOW, end_color=styles.colors.YELLOW)
    greenFill = styles.PatternFill(fill_type='solid', start_color=styles.colors.GREEN, end_color=styles.colors.GREEN)
    redFill = styles.PatternFill(fill_type='solid', start_color=styles.colors.RED, end_color=styles.colors.RED)
    if type_formatting == 'Yellow':
        ws1['{}'.format(case_finding(obj, ws1))].fill = yellowFill
        ws2['{}'.format(case_finding(obj, ws2))].fill = yellowFill
    elif type_formatting == 'Green':
        ws1['{}'.format(case_finding(obj, ws1))].fill = greenFill
        ws2['{}'.format(case_finding(obj, ws2))].fill = greenFill
    elif type_formatting == 'Red':
        ws1['{}'.format(case_finding(obj, ws1))].fill = redFill
        ws2['{}'.format(case_finding(obj, ws2))].fill = redFill
    wbf1.save('test1.xlsx')
    wbf2.save('test2.xlsx')


def write_to_file(filename, lst):
    with open(filename, 'w') as f:
        for item in lst:
            f.write(item[-1])
            f.write('  :  ')
            f.write(item[0])
            f.write('  :  ')
            f.write(item[1])
            f.write('\n')


def compare():
    for obj in second_file:
        if obj[0] not in names_file1:  # Если позиция из нового файла не нашлась в старом
            comparedYellow.append(obj)
            continue
        for obj1 in first_file:
            if obj[0] == obj1[0]:  # Если совпадение нашлось
                if obj[1] == obj1[1]:  # Если отклонения одинаковые
                    comparedGreen.append(obj)
                else:
                    comparedRed.append(obj)


def make_list(lst, count, name, offset):
    for i in range(2, len(name)):
        try:
            lst.append((count[i], name[i], offset[i]))
        except Exception as e:
            messagebox.showinfo('Error', 'iter: #{}\nError: {}'.format(i, e.args))


def main():
    download_info(path1, names_file1, offset_file1, count1)
    download_info(path2, names_file2, offset_file2, count2)
    # print(count1, names_file1, offset_file1, len(count1), len(names_file1), len(offset_file1), sep='\n')
    # print(count2, names_file2, offset_file2, len(count2), len(names_file2), len(offset_file2), sep='\n')
    make_list(first_file, names_file1, offset_file1, count1)
    make_list(second_file, names_file2, offset_file2, count2)
    compare()
    print(len(comparedYellow), len(comparedGreen), len(comparedRed), sep='\n\n')
    write_to_file(out_path + '\\Yellow.txt', comparedYellow)
    write_to_file(out_path + '\\Green.txt', comparedGreen)
    write_to_file(out_path + '\\Red.txt', comparedRed)
    messagebox.showinfo('Done!', 'Работа выполнена. Проверьте указанную папку.')
    clear_vars()


def input1():
    global path1
    path1 = fd.askopenfilename(
        filetypes=(('Another Excel Files', '*.xlsx'), ('Excel Files', '*.xls'), ('All Files', '*.*')))
    first_file_label = tkinter.Label(text=os.path.basename(path1)).place(x=200, y=50)


def input2():
    global path2
    path2 = fd.askopenfilename(
        filetypes=(('Another Excel Files', '*.xlsx'), ('Excel Files', '*.xls'), ('All Files', '*.*')))
    second_file_label = tkinter.Label(text=os.path.basename(path2)).place(x=250, y=100)


def output():
    global out_path
    out_path = fd.askdirectory()
    out_path_label = tkinter.Label(text=out_path).place(x=30, y=178)


if __name__ == '__main__':
    root = tkinter.Tk()
    root.title('Excel Files Comparator')
    root.geometry('400x300')
    clear_vars()
    b1 = tkinter.Button(text='Выберите СТАРЫЙ файл', command=input1).place(x=30, y=50)
    b2 = tkinter.Button(text='Выберите НОВЫЙ(текущий) файл', command=input2).place(x=30, y=100)
    b3 = tkinter.Button(text='Выберите место, куда сохранять результаты', command=output).place(x=30, y=150)
    b4 = tkinter.Button(text='Начать обработку', command=main).place(x=30, y=200)
    root.mainloop()
